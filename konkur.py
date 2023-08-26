from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By


class bcolors:
    ENDC = "\033[0m"
    CRED = "\33[31m"


# your exel file path
EXEL_FILE_PATH = r"C:\Users\Mmd\Desktop\BigProject\finalchoices.xlsx"

wb2 = load_workbook(EXEL_FILE_PATH)

# your exel sheet name(default is Sheet1)
worksheet1 = wb2["Sheet1"]

choices = []
for row in worksheet1.iter_rows():
    choices.append(row[0].value)


URL = "https://my.sanjesh.org/field-selection"

# you can use another browsers as well
driver = webdriver.Edge()
driver.get(URL)

test = input(
    f"{bcolors.CRED}be account khod login konid va be safhe entakhab reshte berid(safheiy ke bayad code reshte haye mokhtalf ra vared konid) va dar nahayat dar terminal kelid ENTER ra bezanid{bcolors.ENDC}\n"
)

for choiceNumers in range(len(choices) - 1):
    add_button = driver.find_element(By.CSS_SELECTOR, 'button[aria-label="add-new"]')
    add_button.click()

numOfUni = 0
for input_uni_number in driver.find_elements(
    By.CSS_SELECTOR, 'input[class="form-control"]'
):
    input_uni_number.send_keys(str(choices[numOfUni]))
    numOfUni += 1

test = input(
    f"{bcolors.CRED}taeid nahaei ro anjam bedid dar nahayat dar terminal kelid ENTER ra bezanid(IMPORTANT: in kar safhe ra mibandad, az save shodan motmaen bashid){bcolors.ENDC}\n"
)

driver.quit()
